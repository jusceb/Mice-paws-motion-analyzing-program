# -*- coding: utf-8 -*-

import sys
import math

from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QApplication, QDialog
from PyQt5.uic import loadUi
import scipy.io


class Program(QDialog):
    def __init__(self):
        super(Program,self).__init__()
        loadUi('mouse_data_gui.ui',self)
        self.setWindowTitle('Analiza ruchu łapek myszy')
        self.pushButton.clicked.connect(self.on_pushButton_clicked)
        self.pushButton_2.clicked.connect(self.clear_pushButton_clicked) 
        self.pushButton_3.clicked.connect(self.close_pushButton_clicked) 

    @pyqtSlot()
    def on_pushButton_clicked(self):

        import os
        import openpyxl
        
        kali = scipy.io.loadmat(self.lineEdit_3.text())
        nazwy = []
        wartosci = []
        for i in kali['kalibracja'][0][0]:
            nazwy.append(str(i[0][0]) + '_zaczepienie')
            for i in kali['kalibracja'][0][1]:
                wartosci.append(i[0])
        liczba_krokow = []
        dystans =[]
        czas_1=[]
        proba = []
        nr_myszy = []
        grupa = []
        tydzien = []
        
        os.chdir(self.lineEdit.text())
        for file in os.listdir(self.lineEdit.text()):            
            if os.path.splitext(file)[1] == '.mat' and os.path.splitext(file)[0] != 'kalibracja':
                temp_kroki = []
                temp_dystans = []
                temp_czas = []
                proba.append(os.path.splitext(file)[0].split()[0])
                grupa.append(os.path.splitext(file)[0].split()[1])
                nr_myszy.append(os.path.splitext(file)[0].split()[5])
                tydzien.append(os.path.splitext(file)[0].split()[2])
                FLx = []
                FLy = []
                FRx = []
                FRy = []
                HLx = []
                HLy = []
                HRx = []
                HRy = []

                mat = scipy.io.loadmat(file)
                a = sorted(mat.keys())
                data = mat['lapki']
                for i in range(len(data)):
                            FLx.append(data[i][0])
                            FLy.append(data[i][1])
                            FRx.append(data[i][2])
                            FRy.append(data[i][3])
                            HLx.append(data[i][4])
                            HLy.append(data[i][5])
                            HRx.append(data[i][6])
                            HRy.append(data[i][7])                            
                kalibracja = 0
                napis = os.path.splitext(file)[0].upper()                
                for i in range(len(nazwy)):
                    if nazwy[i].upper() == napis:
                        kalibracja += float(wartosci[i])
                x = [FLx, FRx, HLx, HRx]
                y = [FLy, FRy, HLy, HRy]
            
                # Liczba kroków
                for j in range(len(x)):                 
                    steps_FL = 0
                    for i in range(len(x[j])-1):
                        if str(x[j][i]) == 'nan' and str(x[j][i+1]) != 'nan':
                            steps_FL += 1
                    temp_kroki.append(steps_FL - 1)
                    
                    #dystans                    
                    a = []
                    for i in range(len(x[j])-1):
                        if str(x[j][i]) == "nan" and str(x[j][i+1]) != 'nan':
                         a.append(float(x[j][i+1]))
                    b = []
                    for i in range(len(y[j])-1):
                        if str(y[j][i]) == "nan" and str(y[j][i+1]) != 'nan':
                         b.append(float(y[j][i+1]))
                
                    dystans_FL = 0
                    for i in range(len(a)-1):
                        krok = math.sqrt((a[i+1] - a[i])**2 + (b[i+1] - b[i])**2)
                        dystans_FL += krok
                    temp_dystans.append(dystans_FL * kalibracja)
                    
                    #czas trwania zaczepienia i odczepienia
                    licz = []
                    for i in range(len(x[j])-1):
                        if str(x[j][i]) == "nan" and str(x[j][i+1]) == "nan":
                            licz.append(1)
                        elif str(x[j][i]) == "nan" and str(x[j][i+1]) != "nan":
                            licz.append(1)
                            licz.append('stop')
                    czas = [] 
                    znacz = []
                    for i in range(len(licz)):
                        if licz[i] == 'stop':
                            znacz.append(i)
                    for i in range(len(znacz)-1):
                        czas.append(sum(licz[znacz[i]+1:znacz[i+1]]))
                    czas_kroku_FL = sum(czas)
                    temp_czas.append(czas_kroku_FL/25)
                liczba_krokow.append(temp_kroki)
                dystans.append(temp_dystans)
                czas_1.append(temp_czas)
            
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.merge_cells('E2:H2')
                sheet.cell(row = 2, column = 5).value = 'Liczba kroków'          
                sheet.merge_cells('I2:L2')
                sheet.cell(row = 2, column = 9).value = 'Dystans'
                sheet.merge_cells('M2:P2')
                sheet.cell(row = 2, column = 13).value = 'Czas trwania kroków'
                sheet.merge_cells('I4:L4')
                sheet.cell(row =4, column = 9).value = '[mm]'
                sheet.merge_cells('M4:P4')
                sheet.cell(row = 4, column = 13).value = '[s]'
                
                lapy = ['FL', 'FR', 'HL', 'HR', 'FL', 'FR', 'HL', 'HR', 'FL', 'FR', 'HL', 'HR']
                etykiety = ['Próba', 'Grupa', 'Nr myszy', 'Tydzień']
                for i, j in enumerate(lapy):
                    sheet.cell(row=3, column=i+5).value = j                 
                for i, j in enumerate(etykiety):
                    sheet.cell(row=4, column=i+1).value = j
                for k in range(len(liczba_krokow)):
                    #print('kkkkk', k)
                    for i, j in enumerate(liczba_krokow[k]):
                        sheet.cell(row=k+5, column=i+5).value = j  
                
                for k in range(len(dystans)):
                    #print('kkkkk', k)
                    for i, j in enumerate(dystans[k]):
                        sheet.cell(row=k+5, column=i+9).value = j  
                        
                for k in range(len(czas_1)):
                    #print('kkkkk', k)
                    for i, j in enumerate(czas_1[k]):
                        sheet.cell(row=k+5, column=i+13).value = j 
                for i,j in enumerate(proba):
                    sheet.cell(row = i+4, column = 1).value = j
                
                for i,j in enumerate(grupa):
                    sheet.cell(row = i+5, column = 2).value = j
                for i,j in enumerate(nr_myszy):
                    sheet.cell(row = i+5, column = 3).value = j
                for i,j in enumerate(tydzien):
                    sheet.cell(row = i+5, column = 4).value = j
                
                from openpyxl.styles import Border, Side
                def set_border(ws, cell_range):
                    thin = Side(border_style="thin", color="000000")
                    for row in ws[cell_range]:
                        for cell in row:
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                
                set_border(sheet, 'E2:P18') 
                set_border(sheet, 'A4:D18') 
                
                wb.save(self.lineEdit_2.text())
                self.label_4.setText("Udało się!")

    @pyqtSlot()
    def clear_pushButton_clicked(self):
        self.lineEdit.clear()
        self.lineEdit_2.clear()
        self.lineEdit_3.clear()
        self.label_4.setText("Wpisz dane ponownie")
        
    @pyqtSlot()
    def close_pushButton_clicked(self):
        self.close()


app = QApplication(sys.argv)
widget = Program()
widget.show()
app.exec_()

#.xlsx